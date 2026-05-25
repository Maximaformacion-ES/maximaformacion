#!/usr/bin/env python3
"""
Import the SEO team's redirect mapping spreadsheet into Strapi
(content-type "redirect").

Sheets consumed:
  - "301s y 410s"  →  245 redirects (status 301) + 13 410s
  - "A revisar"     →  21 redirects mapped but flagged for manual review
  - "301s PDFs"     →  74 PDF redirects (e.g. /wp-content/uploads/... → /recursos/...)

Imágenes sheet is intentionally skipped (separate decision).

Behaviour:
  - Dry-run by default: prints what would be imported, validates rows,
    flags duplicates with what's already in Strapi.
  - With --apply: actually POSTs each row to Strapi.

Environment:
  STRAPI_URL, STRAPI_API_TOKEN must be set (reads ../.env if you don't
  export them).

Usage:
  python3 scripts/import-redirects.py
  python3 scripts/import-redirects.py --apply
"""

import argparse
import os
import re
import sys
import time
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import openpyxl
import urllib.request
import urllib.error
import json


XLSX_PATH = Path(__file__).resolve().parents[2] / "Redirecciones migración - Máxima Formación.xlsx"


def load_env() -> tuple[str, str]:
    """Read STRAPI_URL and STRAPI_API_TOKEN from env or ../.env."""
    url = os.environ.get("STRAPI_URL")
    token = os.environ.get("STRAPI_API_TOKEN")
    if not (url and token):
        env_path = Path(__file__).resolve().parents[1] / ".env"
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, _, v = line.partition("=")
                v = v.strip().strip('"').strip("'")
                if k == "STRAPI_URL" and not url:
                    url = v
                elif k == "STRAPI_API_TOKEN" and not token:
                    token = v
    if not (url and token):
        sys.exit("STRAPI_URL and STRAPI_API_TOKEN must be set (env or .env)")
    return url.rstrip("/"), token


def normalize_path(url_or_path: str) -> Optional[str]:
    """Strip domain, lowercase, drop trailing slash. Mirrors lib/seo/redirects.ts."""
    if not url_or_path:
        return None
    s = url_or_path.strip()
    if not s:
        return None
    # If absolute URL, take the pathname
    if s.startswith("http://") or s.startswith("https://"):
        parsed = urlparse(s)
        s = parsed.path or "/"
    # Drop query/fragment if any
    s = s.split("?")[0].split("#")[0]
    s = s.lower()
    if len(s) > 1 and s.endswith("/"):
        s = s[:-1]
    if not s.startswith("/"):
        s = "/" + s
    return s


def normalize_destination(value: str) -> Optional[str]:
    """
    Destinations can be relative paths or absolute URLs. Keep absolute URLs as-is
    (preserves the canonical host the SEO team chose), normalize relative paths.
    """
    if not value:
        return None
    s = value.strip()
    if not s:
        return None
    if s.startswith("http://") or s.startswith("https://"):
        # Strip trailing slash from URL paths to avoid double-encoding
        return re.sub(r"/$", "", s) if urlparse(s).path != "/" else s
    return normalize_path(s)


def fetch_existing(strapi_url: str, token: str) -> set[str]:
    """Page through /api/redirects collecting all source paths."""
    seen: set[str] = set()
    page = 1
    while True:
        u = (
            f"{strapi_url}/api/redirects"
            f"?pagination%5Bpage%5D={page}&pagination%5BpageSize%5D=100"
            f"&fields%5B0%5D=source"
        )
        req = urllib.request.Request(u, headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        rows = data.get("data") or []
        if not rows:
            break
        for r in rows:
            src = r.get("source") if isinstance(r, dict) else None
            if src:
                seen.add(normalize_path(src) or "")
        meta = data.get("meta", {}).get("pagination", {})
        if page >= meta.get("pageCount", 1):
            break
        page += 1
    return seen


def collect_rules() -> list[dict]:
    """Parse the XLSX and return normalized rules ready for Strapi."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    out: list[dict] = []
    skipped: list[str] = []

    def emit(source, destination, status, note=""):
        src = normalize_path(source) if source else None
        if not src:
            skipped.append(f"empty source: {source!r}")
            return
        if status != 410:
            dst = normalize_destination(destination) if destination else None
            if not dst:
                skipped.append(f"empty destination for {src} (status {status})")
                return
        else:
            dst = ""
        out.append(
            {
                "source": src,
                "destination": dst,
                "statusCode": status,
                "active": True,
                "notes": note or None,
            }
        )

    # Sheet "301s y 410s"
    ws = wb["301s y 410s"]
    for row in ws.iter_rows(values_only=True, min_row=2):
        old, _vercel, new, tipo = row
        if not old:
            continue
        tipo_str = (tipo or "").strip()
        if tipo_str.startswith("410"):
            emit(old, None, 410, "from-sheet: 301s y 410s (410)")
        else:
            emit(old, new, 301, "from-sheet: 301s y 410s")

    # Sheet "A revisar" — applied as-is per ops decision
    ws = wb["A revisar"]
    for row in ws.iter_rows(values_only=True, min_row=2):
        if len(row) < 3:
            continue
        old, _vercel, new = row[0], row[1], row[2]
        if not old:
            continue
        # One row in this sheet uses statusCode 410 in the third column instead
        # of a destination URL. Detect both shapes.
        if isinstance(new, (int, float)) and int(new) == 410:
            emit(old, None, 410, "from-sheet: A revisar (410)")
        else:
            emit(old, new, 301, "from-sheet: A revisar")

    # Sheet "301s PDFs"
    ws = wb["301s PDFs"]
    for row in ws.iter_rows(values_only=True, min_row=2):
        if len(row) < 3:
            continue
        pdf_source, _enlace, dest = row[0], row[1], row[2]
        if not pdf_source or not dest:
            continue
        dest_str = str(dest).strip()
        if dest_str.lower().startswith("(pdf"):
            # Placeholder rows like "(PDF no en inventario — revisar)"
            skipped.append(f"PDF placeholder destination for {pdf_source}: {dest_str}")
            continue
        emit(pdf_source, dest_str, 301, "from-sheet: 301s PDFs")

    return out, skipped


def post_rule(strapi_url: str, token: str, rule: dict) -> tuple[int, str]:
    body = json.dumps({"data": rule}).encode("utf-8")
    req = urllib.request.Request(
        f"{strapi_url}/api/redirects",
        data=body,
        method="POST",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.status, resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", errors="replace")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="Actually upload to Strapi")
    ap.add_argument(
        "--limit", type=int, default=0, help="Process only the first N rules (testing)"
    )
    args = ap.parse_args()

    strapi_url, token = load_env()
    print(f"Strapi: {strapi_url}")
    print(f"Excel:  {XLSX_PATH}")
    if not XLSX_PATH.exists():
        sys.exit(f"Excel file not found at {XLSX_PATH}")

    print("Fetching existing redirects from Strapi…")
    existing = fetch_existing(strapi_url, token)
    print(f"  existing in Strapi: {len(existing)}")

    print("\nParsing spreadsheet…")
    rules, skipped = collect_rules()
    print(f"  rows ready:  {len(rules)}")
    print(f"  rows skipped: {len(skipped)}")

    # Deduplicate within the spreadsheet itself (same source appears twice)
    by_source: dict[str, dict] = {}
    intra_dups: list[str] = []
    for r in rules:
        if r["source"] in by_source:
            intra_dups.append(r["source"])
        else:
            by_source[r["source"]] = r
    rules = list(by_source.values())
    if intra_dups:
        print(f"  intra-spreadsheet duplicates (kept first): {len(intra_dups)}")

    fresh = [r for r in rules if r["source"] not in existing]
    dupes = [r for r in rules if r["source"] in existing]
    print(f"\nWould upload (new):  {len(fresh)}")
    print(f"Already in Strapi:   {len(dupes)}")

    # Status breakdown
    by_status: dict[int, int] = {}
    for r in fresh:
        by_status[r["statusCode"]] = by_status.get(r["statusCode"], 0) + 1
    print(f"  by status: {by_status}")

    if skipped:
        print("\nSkipped reasons (first 10):")
        for line in skipped[:10]:
            print(f"  - {line}")

    print("\nSample of rules to upload (first 5):")
    for r in fresh[:5]:
        d = r["destination"] or "(gone)"
        print(f"  {r['statusCode']}  {r['source']}  →  {d}")

    if not args.apply:
        print("\nDry-run only. Re-run with --apply to upload.")
        return 0

    target = fresh if args.limit == 0 else fresh[: args.limit]
    print(f"\nUploading {len(target)} rules to Strapi…")
    ok, fail = 0, 0
    failures: list[tuple[dict, int, str]] = []
    for i, r in enumerate(target, 1):
        status, body = post_rule(strapi_url, token, r)
        if 200 <= status < 300:
            ok += 1
        else:
            fail += 1
            failures.append((r, status, body[:300]))
        if i % 25 == 0:
            print(f"  {i}/{len(target)}  (ok={ok}, fail={fail})")
        # gentle throttling
        time.sleep(0.02)

    print(f"\nDone.  ok={ok}  fail={fail}")
    if failures:
        print("\nFirst failures:")
        for r, st, body in failures[:5]:
            print(f"  [{st}] {r['source']} → {r['destination'] or '(gone)'}")
            print(f"        {body}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
